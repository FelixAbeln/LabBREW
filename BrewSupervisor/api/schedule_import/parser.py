from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from Services._shared.storage_paths import default_measurements_dir
from Services._shared.wait_engine.parser import (
    normalize_scalar as _shared_normalize_scalar,
)
from Services._shared.wait_engine.parser import (
    parse_condition_expr as _shared_parse_condition_expr,
)
from Services._shared.wait_engine.parser import (
    parse_elapsed_expr as _shared_parse_elapsed_expr,
)
from Services._shared.wait_engine.parser import (
    parse_wait_expr_string as _shared_parse_wait_expr_string,
)
from Services._shared.wait_engine.parser import (
    split_top_level as _shared_split_top_level,
)

DEFAULT_MEASUREMENTS_DIR = default_measurements_dir()


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _cell_bool(value: Any, default: bool = True) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y"}:
        return True
    if text in {"0", "false", "no", "n"}:
        return False
    return default


def _cell_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _cell_list(value: Any) -> list[str]:
    text = _cell_str(value)
    if not text:
        return []
    return [part.strip() for part in text.split(",") if part and part.strip()]


def _read_meta_sheet(wb) -> dict[str, str]:
    if "meta" not in wb.sheetnames:
        raise ValueError("Workbook must contain a 'meta' sheet")
    ws = wb["meta"]
    meta: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key, value = row[:2]
        if key is None:
            continue
        meta[str(key).strip()] = "" if value is None else str(value).strip()
    return meta


def _read_selection_sheet(wb, sheet_name: str) -> list[str]:
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    values: list[str] = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        raw = row[0] if row else None
        text = _cell_str(raw)
        if not text:
            continue
        if text.lower() in {"parameter", "parameters", "name"}:
            continue
        values.append(text)

    seen: set[str] = set()
    deduped: list[str] = []
    for item in values:
        if item in seen:
            continue
        seen.add(item)
        deduped.append(item)
    return deduped


def _meta_get(meta: dict[str, str], *keys: str) -> str | None:
    lower = {str(k).strip().lower(): v for k, v in meta.items()}
    for key in keys:
        value = lower.get(key.lower())
        if value is not None and str(value).strip() != "":
            return str(value).strip()
    return None


def _build_meta_defaults(meta: dict[str, str], wb=None) -> dict[str, Any]:
    measurement_hz_text = _meta_get(
        meta,
        "measurement_hz",
        "measurement.hz",
        "measurement hz",
        "hz",
    )
    try:
        measurement_hz = (
            float(measurement_hz_text) if measurement_hz_text is not None else 10.0
        )
    except (ValueError, TypeError):
        measurement_hz = 10.0

    # Read parameters from 'data' sheet if present
    parameters = []
    if wb is not None:
        parameters = _read_selection_sheet(wb, "data")

    measurement_config = {
        "hz": measurement_hz,
        "output_dir": _meta_get(
            meta, "measurement_output_dir", "measurement.output_dir"
        )
        or DEFAULT_MEASUREMENTS_DIR,
        "output_format": _meta_get(
            meta, "measurement_output_format", "measurement.output_format"
        )
        or "parquet",
        "session_name": _meta_get(
            meta, "measurement_name", "measurement.name", "measurement.session_name"
        ),
    }

    loadstep_duration_text = _meta_get(
        meta,
        "loadstep_duration_seconds",
        "loadstep.duration_seconds",
        "loadstep_default_duration_seconds",
    )
    try:
        loadstep_duration_seconds = (
            float(loadstep_duration_text)
            if loadstep_duration_text is not None
            else 30.0
        )
    except (ValueError, TypeError):
        loadstep_duration_seconds = 30.0
    if loadstep_duration_seconds <= 0:
        loadstep_duration_seconds = 30.0

    measurement_config["loadstep_duration_seconds"] = loadstep_duration_seconds

    if parameters:
        measurement_config["parameters"] = parameters

    package_defaults = {
        "id": _meta_get(meta, "package_id", "package.id", "id"),
        "name": _meta_get(meta, "package_name", "package.name", "name"),
        "version": _meta_get(meta, "package_version", "package.version", "version"),
        "description": _meta_get(
            meta,
            "package_description",
            "package.description",
            "description",
        ),
        "tags": _cell_list(_meta_get(meta, "package_tags", "package.tags", "tags")),
    }

    return {
        "measurement": measurement_config,
        "package": package_defaults,
    }


def _read_steps_sheet(wb, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [
        str(v).strip() if v is not None else ""
        for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        if not any(v is not None and str(v).strip() != "" for v in row.values()):
            continue
        rows.append(row)
    return rows


def _split_top_level(text: str, delimiter: str = ";") -> list[str]:
    return _shared_split_top_level(text, delimiter)


def _normalize_number(value_text: str) -> Any:
    return _shared_normalize_scalar(value_text)


def _parse_actions(cell_value: Any) -> list[dict[str, Any]]:
    text = _cell_str(cell_value)
    if not text:
        return []

    actions: list[dict[str, Any]] = []
    for token in _split_top_level(text, ";"):
        parts = [part.strip() for part in token.split(":")]
        if len(parts) not in {2, 3}:
            raise ValueError(
                f"Invalid action syntax '{token}'. Use target:value[:ramp_seconds]"
            )
        target = parts[0]
        value = _normalize_number(parts[1])
        if len(parts) == 2:
            actions.append(
                {
                    "kind": "write",
                    "target": target,
                    "value": value,
                    "duration_s": None,
                    "owner": None,
                    "params": {},
                }
            )
        else:
            actions.append(
                {
                    "kind": "ramp",
                    "target": target,
                    "value": value,
                    "duration_s": float(parts[2]),
                    "owner": None,
                    "params": {},
                }
            )
    return actions


def _parse_condition(expr: str) -> dict[str, Any]:
    return _shared_parse_condition_expr(expr)


def _parse_elapsed(expr: str) -> dict[str, Any]:
    return _shared_parse_elapsed_expr(expr)


def _parse_wait_expr(expr: str) -> dict[str, Any]:
    return _shared_parse_wait_expr_string(expr)


def _build_step(
    row: dict[str, Any],
    *,
    defaults: dict[str, Any],
) -> dict[str, Any]:
    actions = _parse_actions(row.get("actions"))

    # take_loadstep: blank / 0 = no loadstep; any positive number = duration in seconds.
    # Capture happens after the step's wait condition is met (before_next timing).
    take_loadstep_raw = row.get("take_loadstep")
    take_loadstep_text = _cell_str(take_loadstep_raw)
    take_loadstep_seconds = _cell_float(take_loadstep_raw)

    if take_loadstep_seconds is not None and take_loadstep_seconds > 0:
        actions.append(
            {
                "kind": "take_loadstep",
                "target": None,
                "value": None,
                "duration_s": take_loadstep_seconds,
                "owner": None,
                "params": {"timing": "before_next"},
            }
        )
    elif take_loadstep_text is not None and take_loadstep_seconds is None:
        trigger_wait = _parse_wait_expr(take_loadstep_text)
        if trigger_wait is None:
            raise ValueError(
                f"Invalid take_loadstep value {take_loadstep_raw!r}: "
                "expected numeric seconds or a wait expression"
            )
        default_loadstep_seconds = float(
            defaults.get("measurement", {}).get("loadstep_duration_seconds", 30.0)
            or 30.0
        )
        actions.append(
            {
                "kind": "take_loadstep",
                "target": None,
                "value": None,
                "duration_s": default_loadstep_seconds,
                "owner": None,
                "params": {
                    "timing": "on_trigger",
                    "trigger_wait": trigger_wait,
                },
            }
        )

    return {
        "id": _cell_str(row.get("step_id")),
        "name": _cell_str(row.get("name")),
        "actions": actions,
        "wait": _parse_wait_expr(_cell_str(row.get("wait")) or ""),
        "enabled": _cell_bool(row.get("enabled"), True),
    }


def parse_schedule_workbook(
    file_bytes: bytes, filename: str = "schedule.xlsx"
) -> dict[str, Any]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    meta = _read_meta_sheet(wb)
    defaults = _build_meta_defaults(meta, wb)
    setup_rows = _read_steps_sheet(wb, "setup_steps")
    plan_rows = _read_steps_sheet(wb, "plan_steps")

    return {
        "id": meta.get("id") or filename.rsplit(".", 1)[0],
        "name": meta.get("name") or meta.get("id") or filename,
        "measurement_config": defaults["measurement"],
        "package_defaults": defaults.get("package", {}),
        "workbook_meta": dict(meta),
        "setup_steps": [
            _build_step(
                row,
                defaults=defaults,
            )
            for row in setup_rows
        ],
        "plan_steps": [
            _build_step(
                row,
                defaults=defaults,
            )
            for row in plan_rows
        ],
    }


def collect_workbook_parameter_references(file_bytes: bytes) -> list[dict[str, str]]:
    """Collect workbook-declared parameter references for validation.

    This includes legacy/user-authored selection-list sheets and optional
    per-step parameter columns that may exist in imported workbooks.
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    refs: list[dict[str, str]] = []

    def _push(path: str, source: str, parameter: str) -> None:
        text = _cell_str(parameter)
        if not text:
            return
        refs.append(
            {
                "path": path,
                "source": source,
                "parameter": text,
            }
        )

    m_selection = _read_selection_sheet(wb, "M-SelectionList")
    for idx, item in enumerate(m_selection):
        _push(f"meta.M-SelectionList[{idx}]", "measurement_selection_list", item)

    ls_selection = _read_selection_sheet(wb, "LS-Selection List")
    if not ls_selection:
        ls_selection = _read_selection_sheet(wb, "LS-SelectionList")
    for idx, item in enumerate(ls_selection):
        _push(f"meta.LS-Selection List[{idx}]", "loadstep_selection_list", item)

    for phase in ("setup_steps", "plan_steps"):
        rows = _read_steps_sheet(wb, phase)
        for row_index, row in enumerate(rows):
            measurement_params = _cell_list(row.get("measurement_parameters"))
            for param_index, item in enumerate(measurement_params):
                _push(
                    f"{phase}[{row_index}].measurement_parameters[{param_index}]",
                    "measurement_parameters_column",
                    item,
                )

            loadstep_params = _cell_list(row.get("loadstep_parameters"))
            for param_index, item in enumerate(loadstep_params):
                _push(
                    f"{phase}[{row_index}].loadstep_parameters[{param_index}]",
                    "loadstep_parameters_column",
                    item,
                )

    return refs
