from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from .excel_loader import ExcelScheduleLoader, LoadResult, STEP_HEADERS
from .models import ScheduleStep, StepAction
from .template_exporter import ScheduleTemplateExporter, TemplateExportOptions

SCHEMA_VERSION = 1


@dataclass(slots=True)
class SchedulePayload:
    schema_version: int
    metadata: dict[str, Any]
    source: dict[str, Any]
    startup_steps: list[dict[str, Any]]
    plan_steps: list[dict[str, Any]]

    def to_dict(self) -> dict[str, Any]:
        return {
            "schema_version": self.schema_version,
            "metadata": self.metadata,
            "source": self.source,
            "startup_steps": self.startup_steps,
            "plan_steps": self.plan_steps,
        }


class ScheduleCodec:
    def __init__(self) -> None:
        self.loader = ExcelScheduleLoader()
        self.template_exporter = ScheduleTemplateExporter()

    def validate_payload(self, payload: dict[str, Any]) -> dict[str, Any]:
        if not isinstance(payload, dict):
            raise ValueError("Schedule payload must be a JSON object")

        schema_version = payload.get("schema_version")
        if schema_version is None:
            raise ValueError("Missing schema_version")
        if int(schema_version) != SCHEMA_VERSION:
            raise ValueError(f"Unsupported schema_version: {schema_version}. Expected {SCHEMA_VERSION}")

        metadata = payload.get("metadata")
        source = payload.get("source")
        if metadata is not None and not isinstance(metadata, dict):
            raise ValueError("metadata must be an object")
        if source is not None and not isinstance(source, dict):
            raise ValueError("source must be an object")

        startup_items = payload.get("startup_steps") or []
        plan_items = payload.get("plan_steps") or []
        if not isinstance(startup_items, list) or not isinstance(plan_items, list):
            raise ValueError("startup_steps and plan_steps must be lists")

        startup_steps = [ScheduleStep.from_payload(item, fallback_index=i) for i, item in enumerate(startup_items) if isinstance(item, dict)]
        plan_steps = [ScheduleStep.from_payload(item, fallback_index=i) for i, item in enumerate(plan_items) if isinstance(item, dict)]
        if len(startup_steps) != len(startup_items) or len(plan_steps) != len(plan_items):
            raise ValueError("Each schedule step must be an object")

        issues: list[str] = []
        issues.extend(self._validate_steps(startup_steps, phase_name="startup"))
        issues.extend(self._validate_steps(plan_steps, phase_name="plan"))
        return {
            "ok": not issues,
            "schema_version": SCHEMA_VERSION,
            "message": "Schedule payload is valid" if not issues else "Schedule payload has validation issues",
            "issues": issues,
            "startup_count": len(startup_steps),
            "plan_count": len(plan_steps),
        }


    def _validate_steps(self, steps: list[ScheduleStep], *, phase_name: str) -> list[str]:
        issues: list[str] = []
        seen_indexes: set[int] = set()
        valid_wait_types = {"elapsed_time", "signal", "all_valid"}
        valid_operators = {">=", "<=", ">", "<", "==", "!=", "in_range", "out_of_range", "valid_for"}

        for position, step in enumerate(steps, start=1):
            label = f"{phase_name} step {step.index if step.index is not None else position}"

            if step.index in seen_indexes:
                issues.append(f"{label}: duplicate index {step.index}")
            else:
                seen_indexes.add(step.index)

            if not str(step.name or '').strip():
                issues.append(f"{label}: missing name")

            if not step.actions:
                issues.append(f"{label}: no actions defined")
            else:
                for action_idx, action in enumerate(step.actions, start=1):
                    if not str(action.target_key or '').strip():
                        issues.append(f"{label}: action {action_idx} missing target_key")

            if step.wait_type not in valid_wait_types:
                issues.append(f"{label}: unsupported wait_type '{step.wait_type}'")
                continue

            if step.wait_type == 'elapsed_time':
                if step.duration_s is None or float(step.duration_s) < 0:
                    issues.append(f"{label}: elapsed_time requires duration_s >= 0")

            elif step.wait_type == 'signal':
                if not str(step.wait_source or '').strip():
                    issues.append(f"{label}: signal wait requires wait_source")
                if step.operator not in valid_operators:
                    issues.append(f"{label}: invalid operator '{step.operator}'")
                if step.operator in {'in_range', 'out_of_range'}:
                    if step.threshold_low is None or step.threshold_high is None:
                        issues.append(f"{label}: {step.operator} requires threshold_low and threshold_high")
                elif step.threshold is None:
                    issues.append(f"{label}: signal wait requires threshold")
                if float(step.hold_for_s or 0.0) < 0:
                    issues.append(f"{label}: hold_for_s must be >= 0")

            elif step.wait_type == 'all_valid':
                sources = [s for s in step.valid_sources if str(s).strip()]
                if not sources:
                    issues.append(f"{label}: all_valid requires at least one valid_source")
                if float(step.hold_for_s or 0.0) < 0:
                    issues.append(f"{label}: hold_for_s must be >= 0")

            if step.require_confirmation and not str(step.confirmation_message or '').strip():
                issues.append(f"{label}: confirmation required but confirmation_message is empty")

        return issues

    def load_excel_as_payload(self, workbook_path: str, sheet_name: str = "") -> tuple[LoadResult, dict[str, Any]]:
        result = self.loader.load_workbook(workbook_path, sheet_name)
        payload = self.build_payload(
            startup_steps=result.startup_steps,
            plan_steps=result.plan_steps,
            metadata={
                "workbook_name": Path(workbook_path).name if workbook_path else "",
                "startup_sheet": result.startup_sheet,
                "plan_sheet": result.plan_sheet,
            },
            source={
                "kind": "excel_workbook",
                "name": Path(workbook_path).name if workbook_path else "",
                "path": str(workbook_path or ""),
            },
        )
        return result, payload

    def build_payload(
        self,
        *,
        startup_steps: list[ScheduleStep],
        plan_steps: list[ScheduleStep],
        metadata: dict[str, Any] | None = None,
        source: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        return SchedulePayload(
            schema_version=SCHEMA_VERSION,
            metadata=metadata or {},
            source=source or {},
            startup_steps=[step.to_payload() for step in startup_steps],
            plan_steps=[step.to_payload() for step in plan_steps],
        ).to_dict()

    def parse_payload(self, payload: dict[str, Any]) -> tuple[list[ScheduleStep], list[ScheduleStep], dict[str, Any], dict[str, Any]]:
        validation = self.validate_payload(payload)
        if not validation["ok"]:
            raise ValueError("; ".join(validation["issues"]))
        startup_items = payload.get("startup_steps") or []
        plan_items = payload.get("plan_steps") or []
        startup_steps = [ScheduleStep.from_payload(item, fallback_index=i) for i, item in enumerate(startup_items)]
        plan_steps = [ScheduleStep.from_payload(item, fallback_index=i) for i, item in enumerate(plan_items)]
        metadata = payload.get("metadata") if isinstance(payload.get("metadata"), dict) else {}
        source = payload.get("source") if isinstance(payload.get("source"), dict) else {}
        return startup_steps, plan_steps, metadata, source

    def export_current_workbook(self, path: str | Path, payload: dict[str, Any]) -> Path:
        target = Path(path)
        if target.suffix.lower() != ".xlsx":
            target = target.with_suffix(".xlsx")
        startup_steps, plan_steps, _metadata, _source = self.parse_payload(payload)
        wb = Workbook()
        startup = wb.active
        startup.title = "StartupRoutine"
        startup.append(STEP_HEADERS)
        plan = wb.create_sheet("Plan")
        plan.append(STEP_HEADERS)
        for step in startup_steps:
            startup.append(_step_to_row(step))
        for step in plan_steps:
            plan.append(_step_to_row(step))
        self.template_exporter._style_step_sheet(startup)
        self.template_exporter._style_step_sheet(plan)
        self.template_exporter._add_validations(startup)
        self.template_exporter._add_validations(plan)

        writable = wb.create_sheet("WritableTargets")
        writable.append(["Parameter name", "Friendly name", "Example", "Notes"])
        for row in self.template_exporter._default_writable_targets():
            writable.append(list(row))
        self.template_exporter._style_reference_sheet(writable, {"A": 34, "B": 30, "C": 18, "D": 60})

        signal_sheet = wb.create_sheet("SignalKeyGuide")
        signal_sheet.append(["Signal key", "Friendly name", "Unit"])
        for row in self.template_exporter._default_signal_rows():
            signal_sheet.append(list(row))
        self.template_exporter._style_reference_sheet(signal_sheet, {"A": 34, "B": 30, "C": 12})

        wait_sheet = wb.create_sheet("WaitTypes")
        wait_sheet.append(["Wait type", "Meaning"])
        for row in self.template_exporter.WAIT_TYPES if hasattr(self.template_exporter, 'WAIT_TYPES') else []:
            wait_sheet.append(list(row))
        # Fallback because constants live at module scope.
        if wait_sheet.max_row == 1:
            from .template_exporter import WAIT_TYPES, OPERATORS

            for row in WAIT_TYPES:
                wait_sheet.append(list(row))
            op_sheet = wb.create_sheet("Operators")
            op_sheet.append(["Operator", "Meaning"])
            for row in OPERATORS:
                op_sheet.append(list(row))
        else:
            op_sheet = wb.create_sheet("Operators")
            op_sheet.append(["Operator", "Meaning"])
        if op_sheet.max_row == 1:
            from .template_exporter import OPERATORS

            for row in OPERATORS:
                op_sheet.append(list(row))
        self.template_exporter._style_reference_sheet(wait_sheet, {"A": 18, "B": 92})
        self.template_exporter._style_reference_sheet(op_sheet, {"A": 18, "B": 92})

        howto = wb.create_sheet("HowToUse")
        howto.append(["What to edit", "Details"])
        details = [
            ("StartupRoutine", "Current uploaded startup schedule as currently held by the service."),
            ("Plan", "Current uploaded plan schedule as currently held by the service."),
            ("Export current", "This workbook is a round-trip snapshot. You can edit it locally and upload it back to the service."),
        ]
        for row in details:
            howto.append(list(row))
        self.template_exporter._style_reference_sheet(howto, {"A": 24, "B": 100})

        wb.save(target)
        return target


def _format_threshold(value: Any) -> Any:
    return "" if value is None else value


def _actions_text(actions: list[StepAction]) -> str:
    return ";".join(action.display_text for action in actions)


# Workbook order still uses time_s for duration / hold semantics to preserve compatibility.
def _step_to_row(step: ScheduleStep) -> list[Any]:
    return [
        step.enabled,
        step.index,
        step.name,
        _actions_text(step.actions),
        step.wait_type,
        step.wait_source,
        step.operator,
        _format_threshold(step.threshold),
        _format_threshold(step.threshold_low),
        _format_threshold(step.threshold_high),
        _format_threshold(step.duration_s if step.wait_type == "elapsed_time" else step.hold_for_s),
        ",".join(step.valid_sources),
        step.confirmation_message,
        step.notes,
    ]
