from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

WAIT_TYPES: list[tuple[str, str]] = [
    ("elapsed_time", "Advance after time_s seconds elapse."),
    ("signal", "Advance when wait_source satisfies operator/threshold and stays true for time_s seconds."),
    ("all_valid", "Advance when every valid_sources signal is valid/true for time_s seconds."),
]

OPERATORS: list[tuple[str, str]] = [
    (">=", "Signal is greater than or equal to threshold."),
    ("<=", "Signal is less than or equal to threshold."),
    (">", "Signal is greater than threshold."),
    ("<", "Signal is less than threshold."),
    ("==", "Signal equals threshold."),
    ("!=", "Signal does not equal threshold."),
    ("in_range", "Signal is between threshold_low and threshold_high."),
    ("out_of_range", "Signal is outside threshold_low and threshold_high."),
    ("valid_for", "Used with all_valid to require validity for time_s seconds."),
]

STEP_HEADERS = [
    "enabled",
    "order",
    "step_name",
    "controller_actions",
    "wait_type",
    "wait_source",
    "operator",
    "threshold",
    "threshold_low",
    "threshold_high",
    "time_s",
    "valid_sources",
    "confirmation_message",
    "notes",
]


@dataclass(slots=True)
class TemplateExportOptions:
    include_examples: bool = True
    include_test_routine: bool = True
    workbook_title: str = "FCS Routine Template"
    discovered_signals: list[tuple[str, str, str]] | None = None
    writable_targets: list[tuple[str, str, str, str]] | None = None


class ScheduleTemplateExporter:
    def export(self, path: str | Path, options: TemplateExportOptions | None = None) -> Path:
        options = options or TemplateExportOptions()
        target = Path(path)
        if target.suffix.lower() != ".xlsx":
            target = target.with_suffix(".xlsx")
        wb = Workbook()
        self.populate_workbook(wb, options)
        wb.save(target)
        return target

    def populate_workbook(self, wb: Workbook, options: TemplateExportOptions) -> None:
        startup = wb.active
        startup.title = "StartupRoutine"
        startup.append(STEP_HEADERS)
        plan = wb.create_sheet("Plan")
        plan.append(STEP_HEADERS)

        if options.include_examples:
            for row in self._startup_rows(include_test_routine=options.include_test_routine):
                startup.append(row)
            for row in self._plan_rows(include_test_routine=options.include_test_routine):
                plan.append(row)

        self._style_step_sheet(startup)
        self._style_step_sheet(plan)
        self._add_validations(startup)
        self._add_validations(plan)

        writable = wb.create_sheet("WritableTargets")
        writable.append(["Parameter name", "Friendly name", "Example", "Notes"])
        for row in options.writable_targets or self._default_writable_targets():
            writable.append(list(row))
        self._style_reference_sheet(writable, {"A": 34, "B": 30, "C": 18, "D": 60})

        signal_sheet = wb.create_sheet("SignalKeyGuide")
        signal_sheet.append(["Signal key", "Friendly name", "Unit"])
        for row in options.discovered_signals or self._default_signal_rows():
            signal_sheet.append(list(row))
        self._style_reference_sheet(signal_sheet, {"A": 34, "B": 30, "C": 12})

        wait_sheet = wb.create_sheet("WaitTypes")
        wait_sheet.append(["Wait type", "Meaning"])
        for row in WAIT_TYPES:
            wait_sheet.append(list(row))
        self._style_reference_sheet(wait_sheet, {"A": 18, "B": 92})

        op_sheet = wb.create_sheet("Operators")
        op_sheet.append(["Operator", "Meaning"])
        for row in OPERATORS:
            op_sheet.append(list(row))
        self._style_reference_sheet(op_sheet, {"A": 18, "B": 92})

        howto = wb.create_sheet("HowToUse")
        howto.append(["What to edit", "Details"])
        help_rows = [
            ("StartupRoutine", "Sheet 1 runs before the main plan. Put twin reset, PSU/CAN bring-up, health checks, and startup confirmations here."),
            ("Plan", "Sheet 2 is the actual test or production sequence. Use the same columns as StartupRoutine."),
            ("controller_actions", "Semicolon-separated parameter:value:ramp entries. Use your real ParameterDB names directly, for example set_temp_Fermentor:18; set_pres_Fermentor:0.6:0.01; brewcan.agitator.0.set_pwm:35"),
            ("wait_type", "Choose elapsed_time, signal, or all_valid from the dropdown."),
            ("wait_source", "For signal waits, use a signal key from SignalKeyGuide."),
            ("operator", "Use >=, <=, >, <, ==, !=, in_range, out_of_range, or valid_for."),
            ("threshold / low / high", "Use threshold for single-value comparisons. Use threshold_low and threshold_high for range checks."),
            ("time_s", "For elapsed_time this is the step duration. For signal/all_valid it becomes the hold time before advancing."),
            ("valid_sources", "Comma-separated signals that must be valid/healthy before advancing."),
            ("confirmation_message", "When filled, the service waits for operator confirmation before moving on."),
        ]
        for row in help_rows:
            howto.append(list(row))
        self._style_reference_sheet(howto, {"A": 24, "B": 100})

    def _startup_rows(self, include_test_routine: bool) -> list[list[object]]:
        if not include_test_routine:
            return []
        return [
            [1, 1, "Pulse twin reset on", "twin.reset:1", "elapsed_time", "", "==", "", "", "", 1, "", "", "Start reset pulse"],
            [1, 2, "Pulse twin reset off", "twin.reset:0", "elapsed_time", "", "==", "", "", "", 1, "", "", "End reset pulse"],
            [1, 3, "Enable PSU", "psu.set_voltage:24;psu.set_enable:true", "signal", "psu.connected", "==", True, "", "", 2, "", "", "Wait for PSU connection"],
            [1, 4, "Wait for CAN", "", "signal", "brewcan.connected", "==", True, "", "", 2, "", "", "Communication check"],
            [1, 5, "Validate signals", "", "all_valid", "", "valid_for", "", "", "", 5, "brewcan.temperature.0,brewcan.pressure.0,psu.connected,brewcan.connected", "", "Health gate"],
            [1, 6, "Operator startup check", "", "elapsed_time", "", "==", "", "", "", 0, "", "Confirm vessel, wiring, and media state before plan", "Optional gate"],
        ]

    def _plan_rows(self, include_test_routine: bool) -> list[list[object]]:
        if not include_test_routine:
            return []
        return [
            [1, 1, "Set temperature", "set_temp_Fermentor:18", "signal", "brewcan.temperature.0", ">=", 17.8, "", "", 20, "brewcan.temperature.0", "", "Reach fermentor temperature"],
            [1, 2, "Ramp pressure", "set_pres_Fermentor:0.60:0.01", "signal", "brewcan.pressure.0", ">=", 0.58, "", "", 15, "brewcan.pressure.0", "", "Pressure ramp check"],
            [1, 3, "Start agitator", "brewcan.agitator.0.set_pwm:35", "elapsed_time", "", "==", "", "", "", 60, "", "", "Direct PWM write"],
            [1, 4, "Hold healthy", "", "all_valid", "", "valid_for", "", "", "", 10, "brewcan.temperature.0,brewcan.pressure.0,brewcan.connected,psu.connected", "", "Quality hold"],
            [1, 5, "Operator inspect", "", "elapsed_time", "", "==", "", "", "", 0, "", "Confirm vents, hoses, and trend stability before finish", "Manual gate"],
        ]

    def _style_step_sheet(self, ws) -> None:
        self._style_header_row(ws)
        widths = {
            "A": 10, "B": 10, "C": 28, "D": 56, "E": 16, "F": 34, "G": 14,
            "H": 14, "I": 14, "J": 14, "K": 12, "L": 44, "M": 44, "N": 40,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A2"
        for row_idx in range(2, ws.max_row + 1):
            fill = PatternFill("solid", fgColor="F8FBFF" if row_idx % 2 == 0 else "EEF5FC")
            for cell in ws[row_idx]:
                cell.fill = fill
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    def _style_reference_sheet(self, ws, widths: dict[str, float]) -> None:
        self._style_header_row(ws)
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A2"
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _style_header_row(self, ws) -> None:
        fill = PatternFill("solid", fgColor="1F4E78")
        font = Font(color="FFFFFF", bold=True)
        border = Border(bottom=Side(style="thin", color="D9E2F3"))
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

    def _add_validations(self, ws) -> None:
        op_col = get_column_letter(STEP_HEADERS.index("operator") + 1)
        wait_col = get_column_letter(STEP_HEADERS.index("wait_type") + 1)
        wait_validation = DataValidation(type="list", formula1=f"='WaitTypes'!$A$2:$A${len(WAIT_TYPES)+1}", allow_blank=True)
        op_validation = DataValidation(type="list", formula1=f"='Operators'!$A$2:$A${len(OPERATORS)+1}", allow_blank=True)
        ws.add_data_validation(wait_validation)
        ws.add_data_validation(op_validation)
        wait_validation.add(f"{wait_col}2:{wait_col}500")
        op_validation.add(f"{op_col}2:{op_col}500")

    def _default_signal_rows(self) -> list[tuple[str, str, str]]:
        return [
            ("brewcan.temperature.0", "Fermentor temperature", "°C"),
            ("brewcan.pressure.0", "Fermentor pressure", "bar"),
            ("brewcan.rpm.0", "Agitator RPM", "rpm"),
            ("brewcan.connected", "CAN connected", "bool"),
            ("psu.connected", "PSU connected", "bool"),
        ]

    def _default_writable_targets(self) -> list[tuple[str, str, str, str]]:
        return [
            ("set_temp_Fermentor", "Fermentor temperature setpoint", "18", "Main temperature target"),
            ("set_pres_Fermentor", "Fermentor pressure setpoint", "0.60:0.01", "Target:ramp example"),
            ("brewcan.agitator.0.set_pwm", "Agitator PWM command", "35", "Direct actuator command"),
            ("psu.set_enable", "PSU enable", "true", "Startup command"),
            ("psu.set_voltage", "PSU voltage", "24", "Startup command"),
            ("twin.reset", "Digital twin reset", "1", "Often used as a pulse"),
        ]
