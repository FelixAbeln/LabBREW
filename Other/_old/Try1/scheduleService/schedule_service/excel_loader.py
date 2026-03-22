from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .models import ScheduleStep, StepAction

STEP_HEADERS = [
    "enabled",
    "order",
    "step_name",
    "controller_actions",
    "wait_type",
    "wait_source",
    "operator",
    "threshold",
    "threshold_low",
    "threshold_high",
    "time_s",
    "valid_sources",
    "confirmation_message",
    "notes",
]

WAIT_TYPE_KEYS = {"elapsed_time", "signal", "all_valid", "confirmation"}
OPERATOR_KEYS = {">=", "<=", ">", "<", "==", "!=", "in_range", "out_of_range", "valid_for"}


@dataclass(slots=True)
class LoadResult:
    ok: bool
    message: str
    sheet_names: list[str]
    startup_sheet: str = "StartupRoutine"
    plan_sheet: str = "Plan"
    startup_steps: list[ScheduleStep] = None  # type: ignore[assignment]
    plan_steps: list[ScheduleStep] = None  # type: ignore[assignment]

    def __post_init__(self) -> None:
        if self.startup_steps is None:
            self.startup_steps = []
        if self.plan_steps is None:
            self.plan_steps = []


class ExcelScheduleLoader:
    def load_workbook(self, workbook_path: str, sheet_name: str = "") -> LoadResult:
        path = Path(workbook_path)
        if not path.exists():
            return LoadResult(ok=False, message=f"Workbook not found: {path}", sheet_names=[])

        wb = load_workbook(path, data_only=True)
        sheet_names = list(wb.sheetnames)
        startup_sheet = "StartupRoutine" if "StartupRoutine" in wb.sheetnames else sheet_names[0]
        plan_sheet = sheet_name.strip() if sheet_name.strip() else ("Plan" if "Plan" in wb.sheetnames else sheet_names[-1])

        if startup_sheet not in wb.sheetnames:
            return LoadResult(ok=False, message=f"Startup sheet not found: {startup_sheet}", sheet_names=sheet_names)
        if plan_sheet not in wb.sheetnames:
            return LoadResult(ok=False, message=f"Plan sheet not found: {plan_sheet}", sheet_names=sheet_names)

        startup_steps = self._parse_sheet(wb[startup_sheet])
        plan_steps = self._parse_sheet(wb[plan_sheet])
        message = f"Loaded {len(startup_steps)} startup steps and {len(plan_steps)} plan steps from {path.name}"
        return LoadResult(
            ok=True,
            message=message,
            sheet_names=sheet_names,
            startup_sheet=startup_sheet,
            plan_sheet=plan_sheet,
            startup_steps=startup_steps,
            plan_steps=plan_steps,
        )

    def _parse_sheet(self, ws) -> list[ScheduleStep]:
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_cells:
            return []
        headers = [str(v).strip() if v is not None else "" for v in header_cells]
        header_map = {h.lower(): i for i, h in enumerate(headers) if h}

        def cell(row, name, default=None):
            idx = header_map.get(name.lower())
            if idx is None or idx >= len(row):
                return default
            value = row[idx]
            return default if value is None else value

        steps: list[ScheduleStep] = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=0):
            if not any(v not in (None, "") for v in row):
                continue

            valid_sources = [part.strip() for part in str(cell(row, "valid_sources", "") or "").split(",") if part.strip()]
            wait_type = _normalize_wait_type(cell(row, "wait_type", "elapsed_time"))
            threshold = _parse_threshold_value(cell(row, "threshold"))
            duration = _as_float(cell(row, "duration_s"))
            time_s = _as_float(cell(row, "time_s"))
            hold_for_s = _as_float(cell(row, "hold_for_s"))
            actions = _parse_actions_cell(cell(row, "controller_actions", ""))
            confirmation_message = str(cell(row, "confirmation_message", "") or "").strip()

            step = ScheduleStep(
                index=int(cell(row, "order", cell(row, "index", row_idx)) or row_idx),
                enabled=_as_bool(cell(row, "enabled", True), default=True),
                name=str(cell(row, "step_name", cell(row, "name", f"Step {row_idx + 1}")) or f"Step {row_idx + 1}"),
                wait_type=wait_type,
                wait_source=str(cell(row, "wait_source", "") or "").strip(),
                operator=str(cell(row, "operator", ">=") or ">=").strip().lower(),
                threshold=threshold,
                threshold_low=_as_float(cell(row, "threshold_low")),
                threshold_high=_as_float(cell(row, "threshold_high")),
                duration_s=duration if duration is not None else time_s,
                hold_for_s=max(0.0, hold_for_s if hold_for_s is not None else (time_s or 0.0)),
                valid_sources=valid_sources,
                require_confirmation=bool(confirmation_message),
                confirmation_message=confirmation_message,
                actions=actions,
                notes=str(cell(row, "notes", "") or "").strip(),
            )

            if step.operator not in OPERATOR_KEYS:
                step.operator = ">=" if step.wait_type == "signal" else ("valid_for" if step.wait_type == "all_valid" else step.operator)
            if step.wait_type == "elapsed_time" and step.duration_s is None:
                step.duration_s = max(0.0, float(step.hold_for_s or 0.0))
            elif step.wait_type in {"signal", "all_valid"} and not float(step.hold_for_s or 0.0):
                step.hold_for_s = max(0.0, float(step.duration_s or 0.0))
            if step.wait_type == "confirmation":
                step.wait_type = "elapsed_time"
                step.require_confirmation = True
            steps.append(step)
        return steps


def _as_bool(value, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    txt = str(value).strip().lower()
    if txt in {"1", "true", "yes", "y", "on"}:
        return True
    if txt in {"0", "false", "no", "n", "off"}:
        return False
    return default


def _as_float(value):
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "."))
    except Exception:
        return None


def _parse_optional_bool(value):
    if value in (None, ""):
        return None
    txt = str(value).strip().lower()
    if txt in {"", "none", "keep", "default", "inherit"}:
        return None
    if txt in {"1", "true", "yes", "y", "on"}:
        return True
    if txt in {"0", "false", "no", "n", "off"}:
        return False
    return None


def _normalize_wait_type(value: str) -> str:
    txt = str(value or "elapsed_time").strip().lower()
    alias_map = {
        "signal_threshold": "signal",
        "threshold": "signal",
        "twin_threshold": "signal",
        "valid_all": "all_valid",
        "valid": "all_valid",
        "time": "elapsed_time",
        "duration": "elapsed_time",
    }
    normalized = alias_map.get(txt, txt or "elapsed_time")
    return normalized if normalized in WAIT_TYPE_KEYS else "elapsed_time"


def _parse_scalar_or_text(value_text: str):
    b = _parse_optional_bool(value_text)
    if b is not None:
        return b
    try:
        return float(value_text.replace(",", "."))
    except Exception:
        return value_text


def _parse_threshold_value(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float, bool)):
        return value
    return _parse_scalar_or_text(str(value).strip())


def _parse_action_entry(text: str) -> StepAction | None:
    raw = str(text or "").strip()
    if not raw:
        return None
    parts = [part.strip() for part in raw.split(":")]
    if len(parts) < 2:
        return None

    target_key = parts[0]
    value_text = parts[1]
    ramp = None
    if len(parts) >= 3 and parts[2] not in {"", "-"}:
        try:
            ramp = abs(float(parts[2].replace(",", ".")))
        except Exception:
            ramp = None
    return StepAction(
        target_key=target_key,
        value=_parse_scalar_or_text(value_text),
        ramp_per_s=ramp,
        raw_value=value_text,
    )


def _parse_actions_cell(value) -> list[StepAction]:
    if value in (None, ""):
        return []
    text = str(value).replace("\n", ";")
    actions: list[StepAction] = []
    for chunk in text.split(";"):
        action = _parse_action_entry(chunk)
        if action is not None:
            actions.append(action)
    return actions
