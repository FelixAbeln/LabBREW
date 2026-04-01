from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook

from BrewSupervisor.api.schedule_import import parser, validator


def _wb_bytes(*, include_meta: bool = True) -> bytes:
    wb = Workbook()
    ws = wb.active
    if include_meta:
        ws.title = "meta"
        ws.append(["key", "value"])
    else:
        ws.title = "sheet1"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parser_cell_helpers_and_meta_get() -> None:
    assert parser._cell_str("  alpha  ") == "alpha"
    assert parser._cell_str("   ") is None

    assert parser._cell_bool(None, default=False) is False
    assert parser._cell_bool("yes") is True
    assert parser._cell_bool("0") is False
    assert parser._cell_bool("unknown", default=True) is True

    assert parser._cell_float("2.5") == 2.5
    assert parser._cell_float("x") is None
    assert parser._cell_list("a, b, , c") == ["a", "b", "c"]

    meta = {"Measurement.Hz": "15", "measurement_name": "Batch-1"}
    assert parser._meta_get(meta, "measurement_hz", "measurement.hz") == "15"
    assert parser._meta_get(meta, "missing") is None


def test_read_meta_sheet_requires_meta_tab() -> None:
    wb = Workbook()
    wb.active.title = "other"

    with pytest.raises(ValueError):
        parser._read_meta_sheet(wb)


def test_read_selection_sheet_skips_headers_and_dedupes() -> None:
    wb = Workbook()
    wb.active.title = "meta"
    ws = wb.create_sheet("data")
    ws.append(["Parameter"])
    ws.append(["reactor.temp"])
    ws.append([" reactor.temp "])
    ws.append(["name"])
    ws.append([None])
    ws.append(["reactor.ph"])

    values = parser._read_selection_sheet(wb, "data")
    assert values == ["reactor.temp", "reactor.ph"]
    assert parser._read_selection_sheet(wb, "missing") == []


def test_split_parse_and_wait_helpers_cover_error_paths() -> None:
    assert parser._split_top_level("all(a;b);x") == ["all(a;b)", "x"]
    assert parser._normalize_number("5") == 5
    assert parser._normalize_number("5.5") == 5.5
    assert parser._normalize_number("true") is True
    assert parser._normalize_number("text") == "text"

    assert parser._parse_wait_expr("") is None
    assert parser._parse_wait_expr("elapsed:2") == {"kind": "elapsed", "duration_s": 2.0}

    nested = parser._parse_wait_expr("any(cond:a:>:1;all(elapsed:2;cond:b:<=:3:1))")
    assert nested["kind"] == "any_of"
    assert len(nested["children"]) == 2

    rising = parser._parse_wait_expr("rising(cond:a:>:1)")
    assert rising == {
        "kind": "rising",
        "child": {"kind": "condition", "condition": {"source": "a", "operator": ">", "threshold": 1}},
    }

    pulse = parser._parse_wait_expr("pulse(cond:a:>:1;3)")
    assert pulse["kind"] == "pulse"
    assert pulse["hold_s"] == 3.0

    with pytest.raises(ValueError):
        parser._parse_condition("cond:a:>")
    with pytest.raises(ValueError):
        parser._parse_elapsed("elapsed")
    with pytest.raises(ValueError, match="unmatched closing parenthesis"):
        parser._split_top_level("a);b")
    with pytest.raises(ValueError, match="unmatched opening parenthesis"):
        parser._split_top_level("all(a;b")
    with pytest.raises(ValueError, match="pulse"):
        parser._parse_wait_expr("pulse(cond:a:>:1)")


def test_collect_workbook_references_uses_selection_list_fallback() -> None:
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "meta"
    ws_meta.append(["key", "value"])

    ws_ls = wb.create_sheet("LS-SelectionList")
    ws_ls.append(["name"])
    ws_ls.append(["load.mass"])

    ws_setup = wb.create_sheet("setup_steps")
    ws_setup.append(["step_id", "name", "measurement_parameters", "loadstep_parameters"])
    ws_setup.append(["s1", "Step", "reactor.temp", "load.mass"])

    ws_plan = wb.create_sheet("plan_steps")
    ws_plan.append(["step_id", "name", "measurement_parameters", "loadstep_parameters"])

    buf = BytesIO()
    wb.save(buf)

    refs = parser.collect_workbook_parameter_references(buf.getvalue())
    paths = {item["path"] for item in refs}
    assert "meta.LS-Selection List[0]" in paths
    assert "setup_steps[0].measurement_parameters[0]" in paths
    assert "setup_steps[0].loadstep_parameters[0]" in paths


def test_validator_reports_core_shape_and_parameter_errors() -> None:
    payload = {
        "id": "",
        "name": "",
        "measurement_config": {"parameters": ["unknown.measure"]},
        "setup_steps": "not-a-list",
        "plan_steps": [
            {
                "id": "",
                "name": "",
                "actions": [
                    {"kind": "write", "target": "", "value": None},
                    {"kind": "release_control", "target": "missing.target"},
                    {
                        "kind": "global_measurement",
                        "value": "start",
                        "params": {"parameters": ["missing.param"]},
                    },
                    {
                        "kind": "take_loadstep",
                        "params": {"duration_seconds": "", "parameters": ["unknown.load"]},
                    },
                ],
                "wait": {"kind": "condition", "condition": {"source": "", "operator": "??", "threshold": ""}},
            }
        ],
    }

    result = validator.validate_schedule_payload(payload, available_parameters={"known"})

    assert result["valid"] is False
    assert "MISSING_SCHEDULE_ID" in result["error_codes"]
    assert "MISSING_SCHEDULE_NAME" in result["error_codes"]
    assert "INVALID_PHASE_TYPE" in result["error_codes"]
    assert "UNKNOWN_PARAMETER" in result["error_codes"]
    assert "MISSING_ACTION_TARGET" in result["error_codes"]
    assert "MISSING_WRITE_VALUE" in result["error_codes"]
    assert "MISSING_LOADSTEP_DURATION" in result["error_codes"]
    assert "MISSING_CONDITION_SOURCE" in result["error_codes"]
    assert "INVALID_CONDITION_OPERATOR" in result["error_codes"]
    assert "MISSING_CONDITION_THRESHOLD" in result["error_codes"]


def test_validate_schedule_payload_flags_missing_phase_and_wait_errors() -> None:
    payload = {
        "id": "sched-1",
        "name": "Valid Name",
        "setup_steps": [
            {
                "id": "s1",
                "name": "step",
                "actions": [{"kind": "write", "target": "known", "value": 1}],
                "wait": {"kind": "elapsed", "duration_s": ""},
            }
        ],
    }

    result = validator.validate_schedule_payload(payload, available_parameters={"known"})
    assert result["valid"] is False
    assert "MISSING_PHASE" in result["error_codes"]
    assert "MISSING_ELAPSED_DURATION" in result["error_codes"]


def test_parse_schedule_workbook_raises_without_meta_sheet() -> None:
    with pytest.raises(ValueError):
        parser.parse_schedule_workbook(_wb_bytes(include_meta=False), filename="x.xlsx")


def test_validate_schedule_payload_reports_event_wait_shape_errors() -> None:
    payload = {
        "id": "sched-events",
        "name": "Events",
        "setup_steps": [
            {
                "id": "s1",
                "name": "step",
                "actions": [{"kind": "write", "target": "known", "value": 1}],
                "wait": {"kind": "pulse", "hold_s": "bad", "child": {"kind": "rising"}},
            }
        ],
    }

    result = validator.validate_schedule_payload(payload, available_parameters={"known"})
    codes = set(result["error_codes"])

    assert result["valid"] is False
    assert "MISSING_PHASE" in codes
    assert "INVALID_PULSE_HOLD_SECONDS" in codes
    assert "MISSING_WAIT_CHILD" in codes


def test_validate_schedule_payload_allows_triggered_loadstep_with_measurement_default_duration() -> None:
    payload = {
        "id": "sched-trigger",
        "name": "Trigger",
        "measurement_config": {"loadstep_duration_seconds": 15},
        "setup_steps": [
            {
                "id": "s1",
                "name": "step",
                "actions": [
                    {
                        "kind": "take_loadstep",
                        "params": {
                            "timing": "on_trigger",
                            "trigger_wait": {
                                "kind": "rising",
                                "child": {
                                    "kind": "condition",
                                    "condition": {"source": "known", "operator": "==", "threshold": 1},
                                },
                            },
                        },
                    }
                ],
                "wait": {"kind": "elapsed", "duration_s": 1},
            }
        ],
        "plan_steps": [],
    }

    result = validator.validate_schedule_payload(payload, available_parameters={"known"})
    assert result["valid"] is True
    assert "MISSING_LOADSTEP_DURATION" not in result["error_codes"]
