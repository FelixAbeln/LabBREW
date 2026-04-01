from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook

from BrewSupervisor.api.schedule_import import parser, validator


def _workbook_bytes(
    *,
    meta_rows: list[tuple[object, object]] | None = None,
    setup_rows: list[dict[str, object]] | None = None,
    plan_rows: list[dict[str, object]] | None = None,
    data_rows: list[object] | None = None,
    m_selection_rows: list[object] | None = None,
    ls_selection_rows: list[object] | None = None,
) -> bytes:
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "meta"
    ws_meta.append(["key", "value"])
    for key, value in (meta_rows or []):
        ws_meta.append([key, value])

    if data_rows is not None:
        ws_data = wb.create_sheet("data")
        for item in data_rows:
            ws_data.append([item])

    if m_selection_rows is not None:
        ws_m = wb.create_sheet("M-SelectionList")
        for item in m_selection_rows:
            ws_m.append([item])

    if ls_selection_rows is not None:
        ws_ls = wb.create_sheet("LS-Selection List")
        for item in ls_selection_rows:
            ws_ls.append([item])

    def _append_steps(name: str, rows: list[dict[str, object]] | None) -> None:
        ws = wb.create_sheet(name)
        ws.append([
            "step_id",
            "name",
            "actions",
            "wait",
            "enabled",
            "take_loadstep",
            "measurement_parameters",
            "loadstep_parameters",
        ])
        for row in rows or []:
            ws.append(
                [
                    row.get("step_id"),
                    row.get("name"),
                    row.get("actions"),
                    row.get("wait"),
                    row.get("enabled"),
                    row.get("take_loadstep"),
                    row.get("measurement_parameters"),
                    row.get("loadstep_parameters"),
                ]
            )

    _append_steps("setup_steps", setup_rows)
    _append_steps("plan_steps", plan_rows)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_parse_schedule_workbook_builds_steps_and_measurement_defaults() -> None:
    workbook = _workbook_bytes(
        meta_rows=[
            ("id", "sched-42"),
            ("name", "Lager"),
            ("measurement_hz", "12.5"),
            ("measurement_output_format", "jsonl"),
            ("measurement_name", "batch-a"),
            ("loadstep_duration_seconds", "45"),
        ],
        data_rows=["parameter", "reactor.temp", "pump.speed", "reactor.temp"],
        setup_rows=[
            {
                "step_id": "s1",
                "name": "Heat",
                "actions": "reactor.temp:18:30;pump.enabled:true",
                "wait": "all(elapsed:5;cond:reactor.temp:>=:18:2)",
                "enabled": "yes",
                "take_loadstep": 3,
            }
        ],
        plan_rows=[
            {
                "step_id": "p1",
                "name": "Hold",
                "actions": "reactor.temp:16",
                "wait": "elapsed:120",
                "enabled": "no",
            }
        ],
    )

    payload = parser.parse_schedule_workbook(workbook, filename="lager.xlsx")

    assert payload["id"] == "sched-42"
    assert payload["name"] == "Lager"
    assert payload["measurement_config"] == {
        "hz": 12.5,
        "loadstep_duration_seconds": 45.0,
        "output_dir": "data/measurements",
        "output_format": "jsonl",
        "session_name": "batch-a",
        "parameters": ["reactor.temp", "pump.speed"],
    }

    setup = payload["setup_steps"][0]
    assert setup["id"] == "s1"
    assert setup["enabled"] is True
    assert setup["actions"][0]["kind"] == "ramp"
    assert setup["actions"][1]["kind"] == "write"
    assert setup["actions"][2]["kind"] == "take_loadstep"
    assert setup["actions"][2]["duration_s"] == 3.0
    assert setup["wait"]["kind"] == "all_of"

    plan = payload["plan_steps"][0]
    assert plan["enabled"] is False
    assert plan["wait"] == {"kind": "elapsed", "duration_s": 120.0}


def test_parse_schedule_workbook_uses_filename_and_hz_fallbacks() -> None:
    workbook = _workbook_bytes(meta_rows=[("measurement_hz", "not-a-number")])

    payload = parser.parse_schedule_workbook(workbook, filename="default-name.xlsx")

    assert payload["id"] == "default-name"
    assert payload["name"] == "default-name.xlsx"
    assert payload["measurement_config"]["hz"] == 10.0
    assert payload["measurement_config"]["loadstep_duration_seconds"] == 30.0


def test_parse_schedule_workbook_supports_trigger_wait_in_take_loadstep_column() -> None:
    workbook = _workbook_bytes(
        meta_rows=[("loadstep_duration_seconds", "20")],
        plan_rows=[
            {
                "step_id": "p1",
                "name": "Triggered",
                "actions": "reactor.temp.setpoint:68",
                "take_loadstep": "rising(cond:brew.ready:==:true)",
                "wait": "elapsed:120",
            }
        ],
    )

    payload = parser.parse_schedule_workbook(workbook, filename="triggered.xlsx")
    action = payload["plan_steps"][0]["actions"][1]

    assert action["kind"] == "take_loadstep"
    assert action["duration_s"] == 20.0
    assert action["params"]["timing"] == "on_trigger"
    assert action["params"]["trigger_wait"]["kind"] == "rising"


def test_parse_actions_and_wait_raise_on_invalid_syntax() -> None:
    with pytest.raises(ValueError):
        parser._parse_actions("bad-token")

    with pytest.raises(ValueError):
        parser._parse_wait_expr("unknown(wait)")



def test_collect_workbook_parameter_references_includes_sheet_and_column_entries() -> None:
    workbook = _workbook_bytes(
        m_selection_rows=["parameter", "reactor.temp"],
        ls_selection_rows=["name", "load.mass"],
        setup_rows=[
            {
                "step_id": "s1",
                "name": "Step",
                "measurement_parameters": "reactor.temp, reactor.ph",
                "loadstep_parameters": "load.mass",
            }
        ],
        plan_rows=[
            {
                "step_id": "p1",
                "name": "Step2",
                "measurement_parameters": " reactor.ph ",
            }
        ],
    )

    refs = parser.collect_workbook_parameter_references(workbook)
    ref_paths = {item["path"] for item in refs}
    ref_sources = {item["source"] for item in refs}
    ref_params = {item["parameter"] for item in refs}

    assert "meta.M-SelectionList[0]" in ref_paths
    assert "meta.LS-Selection List[0]" in ref_paths
    assert "setup_steps[0].measurement_parameters[1]" in ref_paths
    assert "setup_steps[0].loadstep_parameters[0]" in ref_paths
    assert "plan_steps[0].measurement_parameters[0]" in ref_paths
    assert "measurement_selection_list" in ref_sources
    assert "measurement_parameters_column" in ref_sources
    assert {"reactor.temp", "reactor.ph", "load.mass"}.issubset(ref_params)


def test_validate_schedule_payload_accepts_warnings_and_collects_unknown_parameter_errors() -> None:
    payload = {
        "id": "sched-1",
        "name": "Schedule",
        "measurement_config": {"parameters": ["known", "unknown.measure"]},
        "setup_steps": [
            {
                "id": "s1",
                "name": "Noop",
                "actions": [],
                "wait": {"kind": "elapsed", "duration_s": 3},
            }
        ],
        "plan_steps": [
            {
                "id": "p1",
                "name": "Write",
                "actions": [
                    {
                        "kind": "write",
                        "target": "known",
                        "value": 1,
                    }
                ],
                "wait": None,
            }
        ],
    }

    result = validator.validate_schedule_payload(
        payload,
        available_parameters={"known"},
        extra_parameter_references=[
            {
                "path": "meta.M-SelectionList[0]",
                "source": "measurement_selection_list",
                "parameter": "unknown.from.workbook",
            }
        ],
    )

    assert result["valid"] is False
    assert "UNKNOWN_PARAMETER" in result["error_codes"]
    assert "STEP_HAS_NO_ACTIONS" in result["warning_codes"]
    assert any(item["path"] == "meta.M-SelectionList[0]" for item in result["issues"])


def test_validate_schedule_payload_reports_action_and_wait_shape_errors() -> None:
    payload = {
        "id": "sched-2",
        "name": "Schedule",
        "setup_steps": [
            {
                "id": "s1",
                "name": "Broken",
                "actions": [
                    {"kind": "invalid"},
                    {"kind": "ramp", "target": "reactor.temp", "value": 15},
                    {"kind": "take_loadstep", "params": {}},
                    {"kind": "global_measurement", "value": "start-now"},
                ],
                "wait": {"kind": "all_of", "children": []},
            }
        ],
        "plan_steps": [
            {
                "id": "p1",
                "name": "BrokenWait",
                "actions": [{"kind": "write", "target": "reactor.temp", "value": 2}],
                "wait": {
                    "kind": "condition",
                    "condition": {
                        "source": "reactor.temp",
                        "operator": "~=",
                        "threshold": "",
                        "for_s": "NaNish",
                    },
                },
            }
        ],
    }

    result = validator.validate_schedule_payload(payload)
    codes = set(result["error_codes"])

    assert result["valid"] is False
    assert "INVALID_ACTION_KIND" in codes
    assert "MISSING_RAMP_DURATION" in codes
    assert "MISSING_LOADSTEP_DURATION" in codes
    assert "INVALID_GLOBAL_MEASUREMENT_MODE" in codes
    assert "MISSING_WAIT_CHILDREN" in codes
    assert "INVALID_CONDITION_OPERATOR" in codes
    assert "MISSING_CONDITION_THRESHOLD" in codes
    assert "INVALID_CONDITION_FOR_SECONDS" in codes
